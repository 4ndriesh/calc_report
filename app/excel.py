from logging_err import Logger
import pandas as pd
from openpyxl import load_workbook, Workbook
import os
from fill_style import Style_pandas
import shutil
from formating import Formating
from copy import copy
from openpyxl.utils import FORMULAE


class Excel():

    def __init__(self):
        self.log = Logger()
        self.log.message_debug("Create class Write_data")
        self.sp = Style_pandas()
        self.fm = Formating()
        self.ido = 1
        # self.dp=Data_preparation()

    def CustomParser(self, data):
        import ast
        res = ast.literal_eval(data)
        print(res)
        # import json
        # n = json.dumps(data)
        # j1 = json.loads(n)
        return res

    def read_xlsx_load_workbook(self, path, file_name, sheet_name='Лист1'):
        wb = load_workbook(filename=os.path.join(path, file_name), read_only=True)
        ws = wb[sheet_name]
        DataFrame = pd.DataFrame()
        for row in ws.iter_rows():
            DataFrame=DataFrame.append(pd.Series([cell.value for cell in row]),ignore_index=True)
        DataFrame = pd.DataFrame(DataFrame.values[1:], columns=DataFrame.iloc[0])
        wb.close()
        return DataFrame

    def read_xlsx_pandas(self, path, file_name, sheet_name='Лист1'):

        try:
            DataFrame = pd.read_excel(os.path.join(path, file_name), sheet_name=sheet_name, engine='openpyxl',
                                      data_only=False)
            # DataFrame=DataFrame.replace('\'','',regex=True)
            # DataFrame.Sum.str.upper()
            # i=DataFrame.Sum.str.replace('_', '',regex=True)
            # DataFrame=DataFrame.dropna(how='all')

        except Exception as e:
            self.log.message_error(e)
            return 0
        return DataFrame

    def create_marks(self, descript, list_mark):
        for row in descript.iter_rows(max_row=22):
            for cell in row:
                mark = str(cell.value).partition("~")[2].partition("~")[0]
                if mark in list_mark:
                    mark_in_excel = {'startcol': cell.column, 'startrow': cell.row}
                    yield mark, mark_in_excel.values()

    def create_xlsx(self, path='', original_file='template.xlsx', copy_file='document.xlsx'):
        shutil.copyfile(os.path.join(path, original_file),

                        os.path.join(path, copy_file))
        main_path = os.path.join(path, copy_file)
        primary = load_workbook(main_path)
        primary.add_named_style(self.fm.template_format())
        primary.add_named_style(self.fm.template_format1())
        primary.save(main_path)
        primary.close()

    def add_sheet_to_xlsx(self, path='', file_name='', new_title=''):
        main_path = os.path.join(path, file_name)
        primary = load_workbook(main_path)
        ws = primary.get_sheet_by_name('ПрайсПО')
        primary.copy_worksheet(ws)
        self.ido += 1
        wss = primary.worksheets[self.ido]
        wss.title = new_title
        primary.save(main_path)
        return primary

    def write_to_excel(self, mv, DataFrame, path='', file_name='document.xlsx', list_columns=None, sheet=''):
        if sheet == '':
            list_sheet = dict(map(lambda ws: (ws.title, ws), mv.worksheets))
        else:
            list_sheet = [sheet]
        for sheet in list_sheet:
            get_descript_sheet = mv[sheet]
            tmp_startrow = 0
            for mark, (startcol, startrow) in self.create_marks(get_descript_sheet, list_columns):
                self.log.message_debug("Finde mark: {}".format(mark))
                if startrow != tmp_startrow and len(DataFrame.index) > 1:
                    tmp_startrow = startrow
                    get_descript_sheet.insert_rows(startrow + 1, len(DataFrame))

                for index, row in enumerate(DataFrame.loc[:, [mark]].iterrows()):
                    if type(get_descript_sheet.cell(row=startrow + index, column=startcol)).__name__ != 'MergedCell':
                        cell = get_descript_sheet.cell(row=startrow, column=startcol)
                        new_cell = get_descript_sheet.cell(row=startrow, column=startcol).offset(row=index, column=0)

                        from openpyxl.formula.translate import Translator
                        from openpyxl.formula import Tokenizer
                        tok = Tokenizer(str(row[1][mark]))
                        if tok.formula.startswith('='):
                            new_cell.value = Translator(row[1][mark], origin=cell.coordinate).translate_formula(
                                row_delta=startrow-2)
                        else:
                            new_cell.value = row[1][mark]

                        if cell.has_style:
                            new_cell.font = copy(cell.font)
                            new_cell.border = copy(cell.border)
                            new_cell.fill = copy(cell.fill)
                            new_cell.number_format = copy(cell.number_format)
                            new_cell.protection = copy(cell.protection)
                            new_cell.alignment = copy(cell.alignment)
                        for key, col, style in self.sp.pars_style_from_dataframe(DataFrame.iloc[index]['Format']):
                            if col == mark:
                                if key == 'style':
                                    get_descript_sheet.cell(row=startrow + index, column=startcol).style = style
                                if key == 'merge_cell':
                                    get_descript_sheet.merge_cells(start_row=startrow + index, start_column=startcol,
                                                                   end_row=startrow + index + style[1],
                                                                   end_column=startcol + style[0])
                                if key == 'hide':
                                    get_descript_sheet.row_dimensions[startrow + index].hidden = True
                                    new_cell.value = style

        mv.save(os.path.join(path, file_name))
        # mv.close()
        self.log.message_debug("Write file: {}".format(path))
