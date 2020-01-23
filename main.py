# -*- coding: utf-8 -*-
__author__ = 'BiziurAA'
from collect_json import Sheets
from read_data import Read_data
from excel import Excel
from pandas import DataFrame
from fill_style import Style_pandas
from data_preparation import Data_preparation
import os
from openpyxl import load_workbook

if __name__ == "__main__":
    s = Sheets()
    r = Read_data()
    w = Excel()
    dp=Data_preparation()

    DataFrame=w.read_xlsx_load_workbook(path='../template', file_name='Data.xlsx',sheet_name='Лист2')
    # DataFrame_title=w.read_xlsx_load_workbook(path='../template', file_name='Data.xlsx',sheet_name='Лист1')
    DataFrame_Data = w.read_xlsx_load_workbook(path='../template', file_name='Data.xlsx', sheet_name='Лист3')



    # columns=list(Data.keys())
    # DataFrame = DataFrame(Data, columns=columns)
    # s = Style_pandas()
    # s.pars_style_from_dataframe(DataFrame)
    # DataFrame=r.read_file_excel()
    distinct_stores = ['document.xlsx']
    reps_per_store = ['one', 'two', 'three']
    w.create_xlsx(path='../template', original_file='template.xlsx', copy_file='document.xlsx')

    DataFrame.at[0, 'Sum'] = 10
    DataFrame.at[1, 'Sum'] = 10
    # DataFrame.at[DataFrame.loc[DataFrame['Name'] == '5. Стоимость ПО ШЛЮЗ Тракт (Чита)'].index, 'Sum'] = 10
    # DataFrame.at[2, 'Sum'] = '=SUM(E12:E13)'
    # DataFrame=DataFrame.drop(DataFrame.loc[DataFrame['Name'] == '5. Стоимость ПО ШЛЮЗ Тракт (Чита)'].index,inplace=True)
    # DataFrame.drop(DataFrame.loc[DataFrame['Name'] == '5. Стоимость ПО ШЛЮЗ Тракт (Чита)'].index,inplace=True)
    # DataFrame['Sum'].loc[(DataFrame['Name'] == '5. Стоимость ПО ШЛЮЗ Тракт (Чита)')]=0
    mv=w.add_sheet_to_xlsx(path='../template', file_name='document.xlsx',new_title='one')
    w.write_to_excel(mv, DataFrame, path='../template', file_name='document.xlsx', list_columns=DataFrame.columns,sheet='one')

    # DataFrame.at[0, 'Sum'] = 10
    # DataFrame.at[1, 'Sum'] = 0
    # mv = w.add_sheet_to_xlsx(path='../template', file_name='document.xlsx', new_title='two')
    # w.write_to_excel(mv, DataFrame, path='../template', file_name='document.xlsx', list_columns=DataFrame.columns,
    #                  sheet='two')
    # w.write_to_excel(mv, DataFrame_title, path='../template', file_name='document.xlsx', list_columns=DataFrame_title.columns)
    # w.write_to_excel(DataFrame, path='../template', list_columns=DataFrame.columns)

    # w.DataFarme_to_Excel(DataFrame)
    # s.create_df()
    # s.write_to_excel()
    # s.read_sheet_by_excel()
    # s.DataFarme_to_Excel()
    # s.DataFarme_to_Json()
